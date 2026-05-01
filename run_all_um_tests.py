#!/usr/bin/env python3
"""
Run all user management test cases and check results.
"""

import subprocess
import sys
import os
from pathlib import Path
import time

def run_user_management_tests():
    """Run all user management tests."""
    print("🚀 Running User Management Test Suite")
    print("=" * 50)

    project_dir = Path(__file__).parent
    os.chdir(project_dir)

    # Run pytest on user management tests
    cmd = [
        sys.executable, "-m", "pytest",
        "tests/test_user_management.py",
        "-v", "--tb=short",
        "--capture=no"
    ]

    print(f"Executing: {' '.join(cmd)}")
    print()

    start_time = time.time()
    try:
        result = subprocess.run(cmd, cwd=project_dir, timeout=600)  # 10 minute timeout
        end_time = time.time()

        print(f"\n✅ Test execution completed in {end_time - start_time:.1f} seconds")
        print(f"Exit code: {result.returncode}")

        return result.returncode

    except subprocess.TimeoutExpired:
        print("\n⏰ Tests timed out after 10 minutes")
        return -1
    except Exception as e:
        print(f"\n❌ Error running tests: {e}")
        return -1

def check_results():
    """Check test results from Excel and other sources."""
    print("\n📊 Checking Test Results")
    print("=" * 30)

    try:
        from openpyxl import load_workbook

        wb = load_workbook('reports/test_cases.xlsx')
        ws = wb.active

        # Get all UM test results
        um_results = []
        for row in range(2, ws.max_row + 1):
            tc_id = ws.cell(row=row, column=1).value
            if tc_id and tc_id.startswith('TC-UM-'):
                status = ws.cell(row=row, column=7).value or '-'
                timestamp = ws.cell(row=row, column=8).value
                um_results.append((tc_id, status, timestamp))

        # Show summary
        total = len(um_results)
        passed = sum(1 for _, status, _ in um_results if status == 'PASS')
        failed = sum(1 for _, status, _ in um_results if status == 'FAIL')
        skipped = sum(1 for _, status, _ in um_results if status == 'SKIP')
        not_run = sum(1 for _, status, _ in um_results if status == '-')

        print(f"Total UM Tests: {total}")
        print(f"✅ Passed: {passed}")
        print(f"❌ Failed: {failed}")
        print(f"⏭️  Skipped: {skipped}")
        print(f"❓ Not Run: {not_run}")

        # Show failed tests
        if failed > 0:
            print(f"\n❌ Failed Tests:")
            for tc_id, status, timestamp in um_results:
                if status == 'FAIL':
                    print(f"   {tc_id} - {timestamp}")

        # Show recent test runs
        print(f"\n🕐 Recent Test Activity:")
        recent_tests = sorted(
            [(tc_id, status, ts) for tc_id, status, ts in um_results if ts],
            key=lambda x: x[2] or '',
            reverse=True
        )[:5]

        for tc_id, status, timestamp in recent_tests:
            print(f"   {tc_id}: {status} - {timestamp}")

    except Exception as e:
        print(f"❌ Error reading Excel results: {e}")

    # Check for bug reports
    bugs_dir = Path('reports/bugs')
    if bugs_dir.exists():
        bug_files = list(bugs_dir.glob('*.json'))
        print(f"\n🐛 Bug Reports: {len(bug_files)}")
        if bug_files:
            # Show recent bug reports
            recent_bugs = sorted(bug_files, key=lambda x: x.stat().st_mtime, reverse=True)[:3]
            for bug_file in recent_bugs:
                print(f"   📄 {bug_file.name}")

    # Check for new screenshots
    screenshots_dir = Path('reports/screenshots')
    if screenshots_dir.exists():
        screenshots = list(screenshots_dir.glob('*.png'))
        print(f"\n📸 Screenshots: {len(screenshots)}")

def main():
    """Main function."""
    # Run the tests
    exit_code = run_user_management_tests()

    # Check results
    check_results()

    print("\n" + "=" * 50)
    print("📋 FINAL SUMMARY:")

    if exit_code == 0:
        print("🎉 ALL TESTS PASSED!")
        print("   ✅ User Management functionality is working correctly")
    elif exit_code == 1:
        print("⚠️  SOME TESTS FAILED!")
        print("   ❌ Check Excel report for failed tests")
        print("   📧 Check email for failure notifications")
        print("   🐛 Check bug reports for details")
    else:
        print("❌ TEST EXECUTION FAILED!")
        print("   Check logs for execution errors")

    print("\n🔍 Results Location:")
    print("   📊 Excel Report: reports/test_cases.xlsx")
    print("   📧 Email Notifications: Check aliantest236@gmail.com")
    print("   🐛 Bug Reports: reports/bugs/")
    print("   📸 Screenshots: reports/screenshots/")
    print("   📝 Logs: reports/logs/dotc_qa_2026-04-28.log")

if __name__ == "__main__":
    main()
