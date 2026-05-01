"""
Excel reporter — creates reports/test_cases.xlsx and updates pass/fail status.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXCEL_PATH = Path("reports/test_cases.xlsx")

ALL_TEST_CASES: list[tuple[str, str, str, str, str, str]] = [
    # (tc_id, module, suite, sub_suite, title, severity)
    ("TC-LOGIN-001", "Login", "Login", "Positive",    "Valid credentials → successful login",              "BLOCKER"),
    ("TC-LOGIN-002", "Login", "Login", "Negative",    "Invalid email → error message",                     "CRITICAL"),
    ("TC-LOGIN-003", "Login", "Login", "Negative",    "Valid email + wrong password → error message",      "CRITICAL"),
    ("TC-LOGIN-004", "Login", "Login", "Validation",  "Empty email → validation error",                   "NORMAL"),
    ("TC-LOGIN-005", "Login", "Login", "Validation",  "Empty password → validation error",                "NORMAL"),
    ("TC-LOGIN-006", "Login", "Login", "Validation",  "Both fields empty → validation error",             "NORMAL"),
    ("TC-LOGIN-007", "Login", "Login", "UI Elements", "Login page elements are visible",                  "NORMAL"),
    ("TC-LOGIN-008", "Login", "Login", "Security",    "SQL injection in email → no crash",                "CRITICAL"),
    ("TC-LOGIN-009", "Login", "Login", "Security",    "XSS in email → sanitized",                        "CRITICAL"),
    ("TC-LOGIN-010", "Login", "Login", "Validation",  "Email format validation in real-time",             "NORMAL"),
    ("TC-LOGIN-011", "Login", "Login", "UI Elements", "Password visibility toggle works",                 "NORMAL"),
    ("TC-LOGIN-012", "Login", "Login", "UI Elements", "Keyboard navigation and Tab support",              "NORMAL"),
    ("TC-LOGIN-013", "Login", "Login", "Validation",  "Form validation clears on user input",             "NORMAL"),
    ("TC-LOGIN-014", "Login", "Login", "Edge Cases",  "Special characters in password accepted",          "NORMAL"),
    ("TC-LOGIN-015", "Login", "Login", "Edge Cases",  "Email whitespace trimming works",                  "MINOR"),
    ("TC-LOGIN-016", "Login", "Login", "Edge Cases",  "Very long email address accepted",                 "MINOR"),
    ("TC-LOGIN-017", "Login", "Login", "Edge Cases",  "Uppercase email handled correctly",                "MINOR"),
    ("TC-LOGIN-018", "Login", "Login", "Accessibility", "Form labels properly associated with inputs",     "NORMAL"),
    ("TC-LOGIN-019", "Login", "Login", "Accessibility", "Page meta description present",                   "MINOR"),
    ("TC-LOGIN-020", "Login", "Login", "Accessibility", "No sensitive data in localStorage/sessionStorage", "CRITICAL"),
    ("TC-UM-001", "User Management", "User Management", "Navigation",   "Navigate to User Management from sidebar",      "BLOCKER"),
    ("TC-UM-002", "User Management", "User Management", "Navigation",   "Page heading is visible",                       "NORMAL"),
    ("TC-UM-003", "User Management", "User Management", "User Table",   "User table is visible",                         "BLOCKER"),
    ("TC-UM-004", "User Management", "User Management", "User Table",   "Table has expected columns",                    "CRITICAL"),
    ("TC-UM-005", "User Management", "User Management", "User Table",   "At least one user record is displayed",         "CRITICAL"),
    ("TC-UM-006", "User Management", "User Management", "User Table",   "Row data fields are non-empty",                 "NORMAL"),
    ("TC-UM-007", "User Management", "User Management", "Filters",      "Status filter → results match",                 "CRITICAL"),
    ("TC-UM-008", "User Management", "User Management", "Filters",      "Location filter → results match",               "CRITICAL"),
    ("TC-UM-009", "User Management", "User Management", "Filters",      "Dating Mode filter → results match",            "CRITICAL"),
    ("TC-UM-010", "User Management", "User Management", "Filters",      "Combined filters → filtered results",           "CRITICAL"),
    ("TC-UM-011", "User Management", "User Management", "Filters",      "Filter producing no results → no-data message", "NORMAL"),
    ("TC-UM-012", "User Management", "User Management", "Filters",      "Reset filters → full list restored",            "NORMAL"),
    ("TC-UM-013", "User Management", "User Management", "Search",       "Search by name → matching results",             "CRITICAL"),
    ("TC-UM-014", "User Management", "User Management", "Search",       "Search by email → matching results",            "CRITICAL"),
    ("TC-UM-015", "User Management", "User Management", "Search",       "Search with no match → no-data message",        "NORMAL"),
    ("TC-UM-016", "User Management", "User Management", "Search",       "Clear search → full list restored",             "NORMAL"),
    ("TC-UM-017", "User Management", "User Management", "User Detail",  "Click View → detail page loads",                "CRITICAL"),
    ("TC-UM-018", "User Management", "User Management", "User Detail",  "Detail page shows user info",                   "CRITICAL"),
    ("TC-UM-019", "User Management", "User Management", "User Detail",  "Navigate back from detail to list",             "NORMAL"),
    ("TC-UM-020", "User Management", "User Management", "Edge Cases",   "Special characters in search → no crash",       "NORMAL"),
    ("TC-UM-021", "User Management", "User Management", "Edge Cases",   "Very long search string → no crash",            "MINOR"),
    ("TC-UM-022", "User Management", "User Management", "Edge Cases",   "Rapid filter switching → stable",               "MINOR"),
    # TC-UM-023 to TC-UM-028 — extended filter tests
    ("TC-UM-023", "User Management", "User Management", "Filters",      "Status filter 'Active' → only Active users shown",       "CRITICAL"),
    ("TC-UM-024", "User Management", "User Management", "Filters",      "Status filter 'Suspended' → only Suspended users shown", "CRITICAL"),
    ("TC-UM-025", "User Management", "User Management", "Filters",      "Status filter 'Suspended' → only Suspended users shown", "CRITICAL"),
    ("TC-UM-026", "User Management", "User Management", "Filters",      "Location filter → each available option → valid results", "NORMAL"),
    ("TC-UM-027", "User Management", "User Management", "Filters",      "Dating Mode filter → each option → valid results",       "NORMAL"),
    ("TC-UM-028", "User Management", "User Management", "Filters",      "Clear All button → resets all filters → full list",      "CRITICAL"),
    # TC-UM-029 to TC-UM-036 — ScoutQA scenario coverage
    ("TC-UM-029", "User Management", "User Management", "Filters",      "Status filter 'All' → full user list shown",             "CRITICAL"),
    ("TC-UM-030", "User Management", "User Management", "Filters",      "Dating Mode 'Active' → only Active dating mode users",   "CRITICAL"),
    ("TC-UM-031", "User Management", "User Management", "Filters",      "Dating Mode 'Inactive' → only Inactive dating mode users", "CRITICAL"),
    ("TC-UM-032", "User Management", "User Management", "Search",       "Search by user ID → matching results",                   "CRITICAL"),
    ("TC-UM-033", "User Management", "User Management", "User Table",   "Table data completeness — name, email, status, location", "CRITICAL"),
    ("TC-UM-034", "User Management", "User Management", "Pagination",   "Pagination is visible on User Management page",          "NORMAL"),
    ("TC-UM-035", "User Management", "User Management", "Pagination",   "Pagination next page navigation works",                  "CRITICAL"),
    ("TC-UM-036", "User Management", "User Management", "Pagination",   "Pagination previous page navigation works",              "CRITICAL"),
    # TC-UM-037 to TC-UM-043 — extended user detail tests
    ("TC-UM-037", "User Management", "User Management", "User Detail",  "Detail page name matches list row data",                  "CRITICAL"),
    ("TC-UM-038", "User Management", "User Management", "User Detail",  "Detail page email matches list row data",                 "CRITICAL"),
    ("TC-UM-039", "User Management", "User Management", "User Detail",  "Detail page status matches list row data",                "CRITICAL"),
    ("TC-UM-040", "User Management", "User Management", "User Detail",  "Detail page URL changes from list URL",                   "CRITICAL"),
    ("TC-UM-041", "User Management", "User Management", "User Detail",  "Admin action buttons visible on user detail page",        "CRITICAL"),
    ("TC-UM-042", "User Management", "User Management", "User Detail",  "Back from detail navigates to user list",                 "NORMAL"),
    ("TC-UM-043", "User Management", "User Management", "User Detail",  "Different users show different detail content",           "NORMAL"),
    # TC-LOGIN-021 to TC-LOGIN-023 — session & logout tests
    ("TC-LOGIN-021", "Login", "Login", "Session", "Login via Enter key on password field",          "CRITICAL"),
    ("TC-LOGIN-022", "Login", "Login", "Session", "Session persists after page refresh",            "CRITICAL"),
    ("TC-LOGIN-023", "Login", "Login", "Session", "Logout redirects back to login page",            "CRITICAL"),
]

HEADERS = ["TC ID", "Module", "Suite", "Sub-suite", "Title", "Severity", "Status", "Last Run"]
COL_WIDTHS = [15, 20, 20, 15, 60, 12, 10, 22]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_STATUS_STYLE: dict[str, tuple[PatternFill, Font]] = {
    "PASS": (PatternFill("solid", fgColor="92D050"), Font(bold=True, color="FFFFFF")),
    "FAIL": (PatternFill("solid", fgColor="FF4C4C"), Font(bold=True, color="FFFFFF")),
    "SKIP": (PatternFill("solid", fgColor="FFC000"), Font(bold=True, color="595959")),
    "-":    (PatternFill("solid", fgColor="F2F2F2"), Font(color="888888")),
}

_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP   = Alignment(wrap_text=True, vertical="center")


class ExcelReporter:
    def __init__(self) -> None:
        self._wb: Workbook | None = None
        self._ws = None
        self._tc_row: dict[str, int] = {}

    def setup(self) -> None:
        EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.title = "Test Results"
        self._ws.row_dimensions[1].height = 30

        for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = self._ws.cell(row=1, column=col, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            self._ws.column_dimensions[get_column_letter(col)].width = width

        for row_idx, (tc_id, module, suite, sub_suite, title, severity) in enumerate(ALL_TEST_CASES, 2):
            row_data = [tc_id, module, suite, sub_suite, title, severity, "-", ""]
            for col, value in enumerate(row_data, 1):
                cell = self._ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = _CENTER if col in (1, 6, 7, 8) else _WRAP

            fill, font = _STATUS_STYLE["-"]
            status_cell = self._ws.cell(row=row_idx, column=7)
            status_cell.fill = fill
            status_cell.font = font

            self._tc_row[tc_id] = row_idx

    def update_result(self, tc_id: str, status: str) -> None:
        row = self._tc_row.get(tc_id)
        if row is None:
            return
        fill, font = _STATUS_STYLE.get(status, _STATUS_STYLE["-"])
        status_cell = self._ws.cell(row=row, column=7, value=status)
        status_cell.fill = fill
        status_cell.font = font
        status_cell.alignment = _CENTER
        self._ws.cell(row=row, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    def save(self) -> None:
        if self._wb:
            try:
                self._wb.save(EXCEL_PATH)
            except PermissionError:
                alt = EXCEL_PATH.with_name(
                    f"test_cases_{datetime.now().strftime('%H%M%S')}.xlsx"
                )
                try:
                    self._wb.save(alt)
                    print(f"[reporter] Excel locked — saved to {alt} instead")
                except Exception:
                    print("[reporter] Could not save Excel report — file may be open")


reporter = ExcelReporter()